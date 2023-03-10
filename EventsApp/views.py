import calendar
import copy
from datetime import datetime, timedelta, date
import time
import openpyxl
import stripe
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
# from hitcount.models import HitCountMixin, HitCount
# from django.contrib.contenttypes.fields import GenericRelation
# from hitcount.views import HitCountDetailView

from Insportify import settings
from .forms import MultiStepForm, AvailabilityForm, LogoForm, InviteForm, NewProfileForm
from .models import master_table, Individual, Organization, Venues, SportsCategory, SportsType, Order, User, \
    Availability, Logo, Extra_Loctaions, Events_PositionInfo, Secondary_SportsChoice, Invite, \
    PositionAndSkillType, SportsImage, Organization_Availability, OrderItems, Advertisement, Profile
import util
from django.db.models import Q
from functools import lru_cache

stripe.api_key = settings.STRIPE_SECRET_KEY

# Cache might be creating run time db issues
# @lru_cache()
def get_profile_from_user(user):
    # user = User.objects.get(user=request.user)
    return Profile.objects.get(active_user=user)


@login_required
def multistep(request):
    entered_values = {}
    sports_type = SportsType.objects.all().order_by('sports_type_text')

    profile = get_profile_from_user(request.user)

    if request.user.is_individual:

        user_loc = Extra_Loctaions.objects.filter(profile=profile).values_list('city', flat=True)
        venues = Venues.objects.values('pk', 'vm_name', 'vm_venuecity').filter(
            vm_venuecity__in=list(user_loc)).order_by('vm_name')
    else:
        venues = Venues.objects.values('pk', 'vm_name').order_by('vm_name')
    if request.method == "POST":
        # Had to remove required since some fieldsets are hidden due to pagination causing client side console errors
        # Checking validity here
        form = MultiStepForm(request.POST)
        # Set entered values
        entered_values = request.POST.dict()
        print(entered_values)
        # Validation
        values_valid = ValidateFormValues(request)
        # print(form.is_valid(), values_valid)
        # Handle Form Post
        if form.is_valid() and values_valid:
            print(list(request.POST.items()))
            # Save data
            obj = form.save(commit=False)
            obj.created_by = profile
            obj.sport_type = request.POST['sport_type']
            obj.venue = request.POST['venue']
            obj.street = request.POST['street']
            obj.city = request.POST['city']
            obj.country = request.POST['country']
            obj.province = request.POST['province']
            obj.zipcode = request.POST['zip_code']
            # obj.position = request.POST['position']
            obj.gender = ','.join(item for item in request.POST.getlist('gender'))
            obj.registration_type = request.POST.get('dropin') if request.POST.get('dropin') is not None else "Drop-in"
            obj.is_recurring = request.POST.get('recurring_event') == "Yes"
            selected_days = request.POST.getlist('recurring_days')
            obj.datetimes_monday = ""
            obj.datetimes_tuesday = ""
            obj.datetimes_wednesday = ""
            obj.datetimes_thursday = ""
            obj.datetimes_friday = ""
            obj.datetimes_saturday = ""
            obj.datetimes_sunday = ""
            obj.datetimes_all = ""
            if obj.is_recurring and 'Monday' in selected_days:
                date = request.POST.get('datetimes_monday_date').split("-")
                start_date = date[0].strip()
                end_date = date[1].strip()
                # print(start_date, end_date)
                ts = datetime.strptime(request.POST.get('datetimes_monday_start_time'), "%H:%M").strftime("%I:%M %p")
                te = datetime.strptime(request.POST.get('datetimes_monday_end_time'), "%H:%M").strftime("%I:%M %p")
                obj.datetimes_monday = start_date + " " + ts + " - " + end_date + " " + te
                obj.datetimes_all += ','.join(days_between(start_date, end_date, "Monday", ts, te)) + ','
            if obj.is_recurring and 'Tuesday' in selected_days:
                date = request.POST.get('datetimes_tuesday_date').split("-")
                start_date = date[0].strip()
                end_date = date[1].strip()
                ts = datetime.strptime(request.POST.get('datetimes_tuesday_start_time'), "%H:%M").strftime("%I:%M %p")
                te = datetime.strptime(request.POST.get('datetimes_tuesday_end_time'), "%H:%M").strftime("%I:%M %p")
                obj.datetimes_tuesday = start_date + " " + ts + " - " + end_date + " " + te
                obj.datetimes_all += ','.join(days_between(start_date, end_date, "Tuesday", ts, te)) + ','
            if obj.is_recurring and 'Wednesday' in selected_days:
                date = request.POST.get('datetimes_wednesday_date').split("-")
                start_date = date[0].strip()
                end_date = date[1].strip()
                ts = datetime.strptime(request.POST.get('datetimes_wednesday_start_time'), "%H:%M").strftime("%I:%M %p")
                te = datetime.strptime(request.POST.get('datetimes_wednesday_end_time'), "%H:%M").strftime("%I:%M %p")
                obj.datetimes_wednesday = start_date + " " + ts + " - " + end_date + " " + te
                obj.datetimes_all += ','.join(days_between(start_date, end_date, "Wednesday", ts, te)) + ','
            if obj.is_recurring and 'Thursday' in selected_days:
                date = request.POST.get('datetimes_thursday_date').split("-")
                start_date = date[0].strip()
                end_date = date[1].strip()
                ts = datetime.strptime(request.POST.get('datetimes_thursday_start_time'), "%H:%M").strftime("%I:%M %p")
                te = datetime.strptime(request.POST.get('datetimes_thursday_end_time'), "%H:%M").strftime("%I:%M %p")
                obj.datetimes_thursday = start_date + " " + ts + " - " + end_date + " " + te
                obj.datetimes_all += ','.join(days_between(start_date, end_date, "Thursday", ts, te)) + ','
            if obj.is_recurring and 'Friday' in selected_days:
                date = request.POST.get('datetimes_friday_date').split("-")
                start_date = date[0].strip()
                end_date = date[1].strip()
                ts = datetime.strptime(request.POST.get('datetimes_friday_start_time'), "%H:%M").strftime("%I:%M %p")
                te = datetime.strptime(request.POST.get('datetimes_friday_end_time'), "%H:%M").strftime("%I:%M %p")
                obj.datetimes_friday = start_date + " " + ts + " - " + end_date + " " + te
                obj.datetimes_all += ','.join(days_between(start_date, end_date, "Friday", ts, te)) + ','
            if obj.is_recurring and 'Saturday' in selected_days:
                date = request.POST.get('datetimes_saturday_date').split("-")
                start_date = date[0].strip()
                end_date = date[1].strip()
                ts = datetime.strptime(request.POST.get('datetimes_saturday_start_time'), "%H:%M").strftime("%I:%M %p")
                te = datetime.strptime(request.POST.get('datetimes_saturday_end_time'), "%H:%M").strftime("%I:%M %p")
                obj.datetimes_saturday = start_date + " " + ts + " - " + end_date + " " + te
                obj.datetimes_all += ','.join(days_between(start_date, end_date, "Saturday", ts, te)) + ','
            if obj.is_recurring and 'Sunday' in selected_days:
                date = request.POST.get('datetimes_sunday_date').split("-")
                start_date = date[0].strip()
                end_date = date[1].strip()
                ts = datetime.strptime(request.POST.get('datetimes_sunday_start_time'), "%H:%M").strftime("%I:%M %p")
                te = datetime.strptime(request.POST.get('datetimes_sunday_end_time'), "%H:%M").strftime("%I:%M %p")
                obj.datetimes_sunday = start_date + " " + ts + " - " + end_date + " " + te
                obj.datetimes_all += ','.join(days_between(start_date, end_date, "Sunday", ts, te)) + ','
            obj.datetimes_exceptions = request.POST.get('datetimes_exceptions') if obj.is_recurring else ""
            obj.datetimes_all = remove_exceptions_from_recurring_days(obj.datetimes_all, obj.datetimes_exceptions)
            obj.datetimes = ""
            if not obj.is_recurring:
                date = request.POST.get('datetimes_date')
                ts = datetime.strptime(request.POST.get('datetimes_start_time'), "%H:%M").strftime("%I:%M %p")
                te = datetime.strptime(request.POST.get('datetimes_end_time'), "%H:%M").strftime("%I:%M %p")
                obj.datetimes = date + " " + ts + " - " + date + " " + te
                obj.datetimes_all = date + " " + ts + " - " + date + " " + te

            obj.save()
            save_event_position_info(request, obj)

            # Take user to event invitation page
            messages.success(request, 'Event Created - Invite Users?')
            redirect_url = f'/invite/' + str(master_table.objects.last().id) + '/'
            return redirect(redirect_url)
        elif not form.is_valid():
            messages.error(request, form.errors)
            messages.error(request, form.non_field_errors)
    else:
        form = MultiStepForm()

    return render(request, 'EventsApp/multi_step.html',
                  {'form': form, 'sports_type': sports_type, 'venues': venues, "values": entered_values})


def ValidateFormValues(request):
    date_valid = True
    event_count_valid = True
    fields_valid = True

    profile = get_profile_from_user(request.user)

    event_count = len(master_table.objects.filter(created_by=profile))
    if not request.user.is_mvp and event_count >= 2:
        messages.error(request, "Cannot create event, maximum events possible by non-MVP member is 2")
        event_count_valid = False
    if not request.POST.get('event_title') or not request.POST.get('venue')\
            or not request.POST.get('event_type') or not request.POST.get('sport_type'):
        messages.error(request, "All fields are required, please enter valid information")
        fields_valid = False
    if not request.POST.get('recurring_event'):
        messages.error(request, "Please select event recurrence")
        date_valid = False
    else:
        if request.POST['recurring_event'] == "Yes":
            selected_days = request.POST.getlist('recurring_days')
            if len(selected_days) < 1:
                messages.error(request, "Please select event days.")
                date_valid = False
            monday_valid = True
            if 'Monday' in selected_days and \
                    (((not request.POST.get('datetimes_monday_date')) or request.POST['datetimes_monday_date'] == "") or
                     ((not request.POST.get('datetimes_monday_start_time')) or request.POST[
                         'datetimes_monday_start_time'] == "") or
                     ((not request.POST.get('datetimes_monday_end_time')) or request.POST[
                         'datetimes_monday_end_time'] == "")):
                messages.error(request, "Invalid date-time selection for Monday, please enter valid dates and times.")
                monday_valid = False
            tuesday_valid = True
            if 'Tuesday' in selected_days and \
                    (((not request.POST.get('datetimes_tuesday_date')) or request.POST[
                        'datetimes_tuesday_date'] == "") or
                     ((not request.POST.get('datetimes_tuesday_start_time')) or request.POST[
                         'datetimes_tuesday_start_time'] == "") or
                     ((not request.POST.get('datetimes_tuesday_end_time')) or request.POST[
                         'datetimes_tuesday_end_time'] == "")):
                messages.error(request, "Invalid date-time selection for Tuesday, please enter valid dates and times.")
                tuesday_valid = False
            wednesday_valid = True
            if 'Wednesday' in selected_days and \
                    (((not request.POST.get('datetimes_wednesday_date')) or request.POST[
                        'datetimes_wednesday_date'] == "") or
                     ((not request.POST.get('datetimes_wednesday_start_time')) or request.POST[
                         'datetimes_wednesday_start_time'] == "") or
                     ((not request.POST.get('datetimes_wednesday_end_time')) or request.POST[
                         'datetimes_wednesday_end_time'] == "")):
                messages.error(request,
                               "Invalid date-time selection for Wednesday, please enter valid dates and times.")
                wednesday_valid = False
            thursday_valid = True
            if 'Thursday' in selected_days and \
                    (((not request.POST.get('datetimes_thursday_date')) or request.POST[
                        'datetimes_thursday_date'] == "") or
                     ((not request.POST.get('datetimes_thursday_start_time')) or request.POST[
                         'datetimes_thursday_start_time'] == "") or
                     ((not request.POST.get('datetimes_thursday_end_time')) or request.POST[
                         'datetimes_thursday_end_time'] == "")):
                messages.error(request, "Invalid date-time selection for Thursday, please enter valid dates and times.")
                thursday_valid = False
            friday_valid = True
            if 'Friday' in selected_days and \
                    (((not request.POST.get('datetimes_friday_date')) or request.POST[
                        'datetimes_friday_date'] == "") or
                     ((not request.POST.get('datetimes_friday_start_time')) or request.POST[
                         'datetimes_friday_start_time'] == "") or
                     ((not request.POST.get('datetimes_friday_end_time')) or request.POST[
                         'datetimes_friday_end_time'] == "")):
                messages.error(request, "Invalid date-time selection for Friday, please enter valid dates and times.")
                friday_valid = False
            saturday_valid = True
            if 'Saturday' in selected_days and \
                    (((not request.POST.get('datetimes_saturday_date')) or request.POST[
                        'datetimes_saturday_date'] == "") or
                     ((not request.POST.get('datetimes_saturday_start_time')) or request.POST[
                         'datetimes_saturday_start_time'] == "") or
                     ((not request.POST.get('datetimes_saturday_end_time')) or request.POST[
                         'datetimes_saturday_end_time'] == "")):
                messages.error(request, "Invalid date-time selection for Saturday, please enter valid dates and times.")
                saturday_valid = False
            sunday_valid = True
            if 'Sunday' in selected_days and \
                    (((not request.POST.get('datetimes_sunday_date')) or request.POST[
                        'datetimes_sunday_date'] == "") or
                     ((not request.POST.get('datetimes_sunday_start_time')) or request.POST[
                         'datetimes_sunday_start_time'] == "") or
                     ((not request.POST.get('datetimes_sunday_end_time')) or request.POST[
                         'datetimes_sunday_end_time'] == "")):
                messages.error(request, "Invalid date-time selection for Sunday, please enter valid dates and times.")
                sunday_valid = False
            date_valid = date_valid and sunday_valid and monday_valid and tuesday_valid and wednesday_valid and thursday_valid \
                         and friday_valid and saturday_valid
        else:
            if ((not request.POST.get('datetimes_date')) or request.POST['datetimes_date'] == "") or \
                    ((not request.POST.get('datetimes_start_time')) or request.POST['datetimes_start_time'] == "") or \
                    ((not request.POST.get('datetimes_end_time')) or request.POST['datetimes_end_time'] == ""):
                messages.error(request, "No date times entered, please enter a date and time for the event")
                date_valid = False
    for i in range(1, 10):
        if request.POST.get('no_of_position' + str(i)) != '' \
                and (request.POST.get('type_of_skill' + str(i)) == ''
                     or request.POST.get('position_cost' + str(i)) == ''
                     or request.POST.get('min_age' + str(i)) == ''
                     or request.POST.get('max_age' + str(i)) == ''):
            messages.error(request, "All position fields are required, please enter valid information")
            fields_valid = False

    return date_valid and event_count_valid and fields_valid


def ValidateUserProfileForm(request):
    valid = True
    print(request.POST)
    if not request.POST.get('first_name') or request.POST['first_name'].strip() == "":
        messages.error(request, "Please enter First Name")
        valid = False
    if not request.POST.get('last_name') or request.POST['last_name'].strip() == "":
        messages.error(request, "Please enter Last Name")
        valid = False
    if not request.POST.get('dob') or request.POST['dob'].strip() == "":
        messages.error(request, "Please enter Date of Birth")
        valid = False

    if not request.POST.get('participation_interest1') and request.POST.get('participation_interest1','').strip() == "" and \
        not request.POST.get('participation_interest2') and request.POST.get('participation_interest2','').strip() == "" and \
        not request.POST.get('participation_interest3') and request.POST.get('participation_interest3','').strip() == "":
        messages.error(request, "Please select Event gender preferences")
        valid = False

    location_valid = False
    if not request.POST.get('city') or request.POST['city'].strip() == "":
        profile = get_profile_from_user(request.user)
        locations = Extra_Loctaions.objects.filter(profile=profile).order_by("city")
        print(locations.all())
        if len(locations) == 0:
            messages.error(request, "Please enter City")
            valid = False
        else:
            location_valid = True
    if location_valid:
        pass
        # TODO: Location Validation code goes here.

    if request.POST.get('province') == "0" or request.POST['province'].strip() == "":
        messages.error(request, "Please enter Province")
        valid = False

    if request.POST.get('country') == "0" or request.POST['country'].strip() == "":
        messages.error(request, "Please enter Country")
        valid = False

    return valid


def ValidateOrgProfileForm(request):
    valid = True
    if not request.POST.get('company_name') or request.POST['company_name'].strip() == "":
        messages.error(request, "Please enter Organization Name")
        valid = False
    if not request.POST.get('registration') or request.POST['registration'].strip() == "":
        messages.error(request, "Please enter Registration Number")
        valid = False
    # if not request.POST.get('email') or request.POST['email'].strip() == "":
    #     messages.error(request, "Please enter Contact Email")
    #     valid = False
    if not request.POST.get('phone') or request.POST['phone'].strip() == "":
        messages.error(request, "Please enter Contact Phone Number")
        valid = False
    # if len(context['locations']) == 0:
    if not request.POST.get('street_name') or request.POST['street_name'].strip() == "":
        messages.error(request, "Please enter Street Name")
        valid = False
    if not request.POST.get('city') or request.POST['city'].strip() == "":
        messages.error(request, "Please enter City")
        valid = False
    if not request.POST.get('province') or request.POST['province'].strip() == "":
        messages.error(request, "Please enter Province")
        valid = False
    if not request.POST.get('country') or request.POST['country'].strip() == "":
        messages.error(request, "Please enter Country")
        valid = False
    if not request.POST.get('postal_code') or request.POST['postal_code'].strip() == "":
        messages.error(request, "Please enter Postal Code")
        valid = False
    if not request.POST.get('gender') or request.POST['gender'].strip() == "":
        messages.error(request, "Please select Focus")
        valid = False
    if not request.POST.get('age_group') or request.POST['age_group'].strip() == "":
        messages.error(request, "Please select Age Group")
        valid = False
    # if not request.POST.get('sport_type') or request.POST['sport_type'].strip() == "":
    #     messages.error(request, "Please select Sports you facilitate")
    #     valid = False

    return valid


def get_venue_details(request):
    data = {}
    if request.method == "POST":
        selected_venue = request.POST['selected_venue']
        try:
            selected_venue = Venues.objects.filter(vm_name=selected_venue).order_by('vm_name')
            # print(selected_venue)
        except Exception:
            data['error_message'] = 'error'
            return JsonResponse(data)
        return JsonResponse(list(selected_venue.values()), safe=False)


def save_event_position_info(request, event):
    if event.datetimes:
        for i in range(1, 10):
            if 'no_of_position' + str(i) in request.POST and request.POST['no_of_position' + str(i)].strip() != "":
                position_name = request.POST['position_name' + str(i)].strip()
                position_type = request.POST['type_of_skill' + str(i)].strip()
                no_of_position = request.POST['no_of_position' + str(i)].strip()
                position_cost = round(float(request.POST['position_cost' + str(i)]), 2)
                min_age = request.POST['min_age' + str(i)].strip()
                max_age = request.POST['max_age' + str(i)].strip() if request.POST['max_age' + str(i)] else "999"
                datetimes = event.datetimes
                # print(position_name, position_type, position_cost, min_age, max_age)
                obj = Events_PositionInfo(event=event, position_name=position_name, max_age=max_age, min_age=min_age,
                                          no_of_position=no_of_position,
                                          position_cost=position_cost, position_number=i,
                                          position_type=position_type, datetimes=datetimes)
                obj.save()
    elif event.datetimes_all:
        all_dates_arr = event.datetimes_all.strip(',').split(',')
        for date in all_dates_arr:
            for i in range(1, 10):
                if 'no_of_position' + str(i) in request.POST and request.POST['no_of_position' + str(i)].strip() != "":
                    position_name = request.POST['position_name' + str(i)].strip()
                    position_type = request.POST['type_of_skill' + str(i)].strip()
                    no_of_position = request.POST['no_of_position' + str(i)].strip()
                    position_cost = round(float(request.POST['position_cost' + str(i)]), 2)
                    min_age = request.POST['min_age' + str(i)].strip()
                    max_age = request.POST['max_age' + str(i)].strip() if request.POST['max_age' + str(i)] else "999"
                    # print(position_name, position_type, position_cost, min_age, max_age)
                    obj = Events_PositionInfo(event=event, position_name=position_name, max_age=max_age,
                                              min_age=min_age,
                                              no_of_position=no_of_position,
                                              position_cost=position_cost, position_number=i,
                                              position_type=position_type, datetimes=date)
                    obj.save()


@login_required
def all_events(request):

    profile = get_profile_from_user(request.user)

    expired_events = []
    event_list = list(master_table.objects.filter(created_by=profile))
    event_list = get_events_by_time(event_list)
    today = date.today()
    for event in event_list[:]:
        datetimes = event.datetimes if event.datetimes else event.current_datetimes
        date_split = datetimes.split(" ")
        datetime_obj = datetime.strptime(date_split[0].strip(), '%m/%d/%Y').date()
        if datetime_obj < today:
            expired_events.append(event)
            event_list.remove(event)

    sort_events_by_date(event_list)
    sort_events_by_date(expired_events)

    event_list = format_time(event_list)

    page_num = request.GET.get('page', 1)
    paginator = Paginator(event_list, 6)

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    return render(request, 'EventsApp/event_list.html', {'event_list': event_list, 'expired_events': expired_events,
                                                         'page_obj': page_obj})


@login_required
def committed_events(request):
    event_list = set()
    profile = get_profile_from_user(request.user)
    cart = OrderItems.objects.filter(profile=profile)
    if len(cart) > 0:
        for item in cart:
            event = master_table.objects.get(pk=item.event.pk)
            event_list.add(event)
        event_list = format_time(event_list)
    return render(request, 'EventsApp/events_committed.html', {'event_list': list(event_list)})


@login_required
def delete_profile(request):
    profile = get_profile_from_user(request.user)
    if profile.is_master:
        User.objects.filter(pk=request.user.pk).delete()
        logout(request)
        return redirect('EventsApp:home')
    else:
        return delete_current_profile(request)


def modify_individual(response, individual):
    if "first_name" in response:
        individual.first_name = response["first_name"].strip() if response["first_name"] else ""
    if "is_current" in response:
        individual.is_current = response["is_current"].strip()
    if "last_name" in response:
        individual.last_name = response["last_name"].strip() if response["last_name"] else ""
    if "phone" in response:
        phone = response["phone"].strip()
        phone = ''.join(i for i in phone if i.isdigit())
        individual.phone = phone
    if "website" in response:
        individual.website = response["website"].strip()
    if "job_title" in response:
        individual.job_title = response["job_title"].strip()
    if "dob" in response:
        individual.dob = response["dob"].strip() if response["dob"] else ""
    if "is_concussion" in response:
        individual.concussion = response["is_concussion"].strip()
    if "is_student" in response:
        individual.is_student = response["is_student"].strip()
    participation_interest = []
    if "participation_interest1" in response:
        participation_interest.append(response['participation_interest1'])
    if "participation_interest2" in response:
        participation_interest.append(response['participation_interest2'])
    if "participation_interest3" in response:
        participation_interest.append(response['participation_interest3'])
    individual.participation_interest = ','.join(item for item in participation_interest)

    if "pronoun" in response:
        individual.pronoun = response["pronoun"].strip() if response["pronoun"] else ""

    if "city" in response:
        individual.city = response["city"].strip() if response["city"] else ""
    if "province" in response:
        individual.province = response["province"].strip() if response["province"] else ""
    if "country" in response:
        individual.country = response["country"].strip() if response["country"] else ""
    if "sport_type" in response:
        individual.sports_type = response["sport_type"].strip() if response["sport_type"] else ""
    if "position" in response:
        individual.sports_position = response["position"].strip() if response["position"] else ""
    if "skill" in response:
        individual.sports_skill = response["skill"].strip() if response["skill"] else ""
    return individual


@login_required
def display_profile(request):
    profiles = Profile.objects.filter(user=request.user)
    context = {
        'profiles_list': profiles
    }
    return render(request, 'EventsApp/switch_profile.html', context)


def _switch_profile(user, name):
    profile = Profile.objects.get(user=user, name=name)
    if not profile:
        print('Profile not found : ', name)
        return False

    curr_prof = get_profile_from_user(user)
    if curr_prof.name == name:
        return True

    curr_prof.active_user = None
    # request.user.profile = None
    curr_prof.save()

    # Assuming name is unique
    profile.active_user = user
    profile.save()

    user.active_profile_name = profile.name
    user.profile_status = profile.profile_status
    user.save()

    # get_profile_from_user.cache_clear()
    return True


@login_required
def switch_profile(request, name):
    _switch_profile(request.user, name)
    return home(request)


@login_required
def delete_current_profile(request):
    curr_prof = get_profile_from_user(request.user)
    master_prof = Profile.objects.get(user=request.user, is_master=True)
    _switch_profile(request.user, master_prof.name)
    if not curr_prof.is_master:
        curr_prof.delete()
    else:
        print('Cannot delete master profile.')
    get_profile_from_user.cache_clear()
    return home(request)


@login_required
def create_profile(request):
    if request.user.is_organization:
        return home(request)

    if request.method == 'GET':
        form = NewProfileForm()
        context = {'form': form}
        return render(request, 'EventsApp/create_profile.html', context)

    if request.method == 'POST':
        form = NewProfileForm(request.POST)
        if form.is_valid():
            # TODO : Add validation
            name = form.cleaned_data.get('name').lower()

            existing_profiles = Profile.objects.filter(user=request.user)
            print(existing_profiles)
            for prof in existing_profiles:
                if prof.name == name:
                    messages.error(request, "A profile with this name already exists. Use another name.")
                    form = NewProfileForm()
                    context = {'form': form}
                    return render(request, 'EventsApp/create_profile.html', context)

            profile = Profile.objects.create(active_user=None, user=request.user, profile_status=False, name=name,
                                             is_master=False)
            profile.save()

            individual = Individual.objects.create(profile=profile, first_name=name)
            individual.save()

            _switch_profile(request.user, name)

            # get_profile_from_user.cache_clear()

            request.user.refresh_from_db()
            # context = {'individual': individual, 'sports_type': [], 'sec_sport_choices': [], 'locations': [], 'user_avaiability': []}
            # return render(request, 'registration/individual_view.html', context)

            return redirect('EventsApp:user_profile')
        else:
            print('Form not valid')
            return render(request, 'EventsApp/create_profile.html')

    print('Error in profile creation.')
    return home(request)


@login_required
def user_profile(request):
    # Displays or updates user profile
    profile = get_profile_from_user(request.user)
    individual = Individual.objects.get(profile=profile)
    context = {
        'individual': individual,
        'profile': profile
    }
    sports_type = SportsType.objects.all().order_by('sports_type_text')
    sec_sport_choices = Secondary_SportsChoice.objects.filter(profile=profile).order_by("sport_type")
    locations = Extra_Loctaions.objects.filter(profile=profile).order_by("city")
    user_avaiability = Availability.objects.filter(profile=profile)
    get_day_of_week(user_avaiability)
    context['sports_type'] = sports_type
    context['sec_sport_choices'] = sec_sport_choices
    context['locations'] = locations
    context['user_avaiability'] = user_avaiability

    return render(request, 'registration/individual_view.html', context)


@login_required
def user_profile_submit(request):
    print('GET request fields', request.GET.dict())
    print('POST request fields', request.POST.dict())

    if not ValidateUserProfileForm(request):
        print('Validation Failed')
        profile = get_profile_from_user(request.user)
        context = {
            'individual': request.POST.dict(),
            'profile': profile
        }
        print(context)
        sports_type = SportsType.objects.all().order_by('sports_type_text')
        sec_sport_choices = Secondary_SportsChoice.objects.filter(profile=profile).order_by("sport_type")
        locations = Extra_Loctaions.objects.filter(profile=profile).order_by("city")
        user_avaiability = Availability.objects.filter(profile=profile)
        get_day_of_week(user_avaiability)
        context['sports_type'] = sports_type
        context['sec_sport_choices'] = sec_sport_choices
        context['locations'] = locations
        context['user_avaiability'] = user_avaiability
        return render(request, 'registration/individual_view.html', context)
    else:

        profile = get_profile_from_user(request.user)
        request.user.profile_status = True
        profile.profile_status = True
        individual = Individual.objects.get(profile=profile)
        individual = modify_individual(request.POST.dict(), individual)
        profile.save()
        request.user.save()
        individual.save()
        print('Individual details updated!')
        return home(request)

    # messages.success(request, 'Individual details updated!')
    # return render(request, 'registration/individual_view.html', context)


def add_sports_positions(request):
    if request.method == "POST":
        selected_sport = request.POST['selected_sport_text']
        selected_position = request.POST['selected_position_text']
        selected_skill = request.POST['selected_skill_text']
        # print(selected_skill, selected_position, selected_sport)

        profile = get_profile_from_user(request.user)

        try:
            if Secondary_SportsChoice.objects.filter(profile=profile, sport_type=selected_sport,
                                                     position=selected_position, skill=selected_skill).exists():
                return JsonResponse({'status': 'Duplicate Position cannot be added!'}, safe=False)
            else:
                obj = Secondary_SportsChoice(profile=profile, sport_type=selected_sport,
                                             position=selected_position, skill=selected_skill)
                obj.save()
            return JsonResponse({'status': 'New Position added!'}, safe=False)
        except Exception:
            return JsonResponse({'status': 'An error occured!'}, safe=False)


def fetch_user_sports_positions(request):
    if request.is_ajax():
        profile = get_profile_from_user(request.user)
        position_choices = Secondary_SportsChoice.objects.filter(profile=profile).order_by("sport_type")
        position_choices = list(position_choices.values("sport_type", "position", "skill", "pk"))
        return JsonResponse(position_choices, safe=False)


def delete_sports_choice(request):
    if request.POST:
        try:
            id = request.POST['id']
            obj = Secondary_SportsChoice.objects.get(pk=id)
            obj.delete()
            return JsonResponse({'status': 'Sports Choice removed successfully!'}, safe=False)
        except:
            return JsonResponse({'status': 'Some error occurred!'}, safe=False)


def add_user_locations(request):
    if request.method == "POST":
        selected_city = request.POST['selected_city_text'].strip() if request.POST['selected_city_text'] else ""
        selected_province = request.POST['selected_province_text'].strip() if request.POST[
            'selected_province_text'] else ""
        selected_country = request.POST['selected_country_text'].strip() if request.POST[
            'selected_country_text'] else ""
        # print(selected_city, selected_province, selected_country)
        try:
            if selected_city != "" and selected_province != "" and selected_country != "":
                profile = get_profile_from_user(request.user)
                if Extra_Loctaions.objects.filter(profile=profile, city=selected_city,
                                                  province=selected_province, country=selected_country).exists():
                    return JsonResponse({'status': 'Duplicate Location cannot be added!'}, safe=False)
                else:
                    obj = Extra_Loctaions(profile=profile, city=selected_city,
                                          province=selected_province, country=selected_country)
                    obj.save()
                return JsonResponse({'status': 'New Location added!'}, safe=False)
            else:
                return JsonResponse({'status': 'Missing values!'}, safe=False)
        except Exception:
            return JsonResponse({'status': 'An error occured!'}, safe=False)


def fetch_user_locations(request):
    if request.is_ajax():
        profile = get_profile_from_user(request.user)
        location_choices = Extra_Loctaions.objects.filter(profile=profile).order_by("city")
        location_choices = list(location_choices.values("city", "province", "country", "pk"))
        return JsonResponse(location_choices, safe=False)


def delete_user_location(request):
    if request.POST:
        try:
            id = request.POST['id']
            # print(id)
            obj = Extra_Loctaions.objects.get(pk=id)
            obj.delete()
            return JsonResponse({'status': 'Location removed successfully!'}, safe=False)
        except:
            return JsonResponse({'status': 'Some error occurred!'}, safe=False)


def get_selected_sports_type(request):
    data = {}
    if request.method == "POST":
        selected_category = request.POST['selected_category_text']
        try:
            selected_type = SportsType.objects.filter(sports_category__sports_catgeory_text=selected_category)
            # print(selected_type)
        except Exception:
            data['error_message'] = 'error'
            return JsonResponse(data)
        return JsonResponse(list(selected_type.values('pk', 'sports_type_text')), safe=False)


def get_sports_type(request):
    data = {}
    if request.method == "GET":
        try:
            sports_type = SportsType.objects.all().order_by('sports_type_text')
        except Exception:
            data['error_message'] = 'error'
            return JsonResponse(data)
        return JsonResponse(list(sports_type.values('pk', 'sports_type_text')), safe=False)


def get_selected_sports_position_and_skill(request):
    data = {}
    data_list = []
    if request.method == "POST":
        selected_sport = request.POST['selected_type_text']
        try:
            positions = PositionAndSkillType.objects.filter(sports_type__sports_type_text=selected_sport).values(
                'position_type').distinct('position_type')
            for position in positions:
                skills = PositionAndSkillType.objects.filter(position_type=position.position_type).values(
                    'pk', 'skill_type').distinct('skill_type')

                print(position.position_type, skills)
            print(data_list)
        except Exception:
            data['error_message'] = 'error'
            return JsonResponse(data)
        return JsonResponse()


def get_selected_sports_skill(request):
    data = {}
    if request.method == "POST":
        sport_position = request.POST['selected_position_text']
        selected_sport = request.POST['selected_type_text']
        # print(selected_sport, sport_position)
        try:
            selected_skills = PositionAndSkillType.objects.filter(sports_type__sports_type_text=selected_sport).filter(
                position_type=sport_position).values('pk', 'skill_type').distinct('skill_type')
            # print(selected_skills)
        except Exception:
            data['error_message'] = 'error'
            return JsonResponse(data)
        return JsonResponse(list(selected_skills.values('pk', 'skill_type')), safe=False)


def get_selected_sports_positions(request):
    data = {}
    if request.method == "POST":
        selected_sport = request.POST['selected_type_text']
        # print(selected_sport)
        try:
            selected_skills = PositionAndSkillType.objects.filter(sports_type__sports_type_text=selected_sport).values(
                'pk', 'position_type').distinct('position_type')
        except Exception:
            data['error_message'] = 'error'
            return JsonResponse(data)
        return JsonResponse(list(selected_skills.values('pk', 'position_type')), safe=False)


@login_required
def organization_profile(request):
    profile = get_profile_from_user(request.user)
    context = {
        'profile': profile
    }
    if request.method == "GET":
        organization = Organization.objects.get(profile=profile)
        # print(organization.__dict__)
        locations = Extra_Loctaions.objects.filter(profile=profile).order_by("city")
        sports_type = SportsType.objects.all().order_by('sports_type_text')
        context['locations'] = locations
        context['sports_type'] = sports_type
        context['organization'] = organization
        return render(request, 'registration/organization_view.html', context)

    if request.method == "POST":
        organization = Organization.objects.filter(profile=profile)
        locations = Extra_Loctaions.objects.filter(profile=profile).order_by("city")
        sports_type = SportsType.objects.all().order_by('sports_type_text')
        context['organization'] = organization
        context['locations'] = locations
        context['sports_type'] = sports_type
        response = request.POST.dict()
        if not ValidateOrgProfileForm(request):
            organization = Organization.objects.get(profile=profile)
            context['organization'] = request.POST.dict()
            return render(request, 'registration/organization_view.html', context)
        else:  # organization.exists():
            organization = Organization.objects.get(profile=profile)
            organization.profile = profile
            if "type_of_organization" in response:
                organization.type_of_organization = response["type_of_organization"].strip() if response[
                    "type_of_organization"] else ""
            if "company_name" in response:
                organization.organization_name = response["company_name"].strip() if response["company_name"] else ""
            if "parent_organization" in response:
                organization.parent_organization_name = response["parent_organization"].strip() if response[
                    "parent_organization"] else ""
            if "registration" in response:
                organization.registration_no = response["registration"].strip() if response["registration"] else ""
            if "year_established" in response:
                organization.year_established = response["year_established"].strip() if response[
                    "year_established"] else ""
            if "street_name" in response:
                organization.street = response["street_name"].strip() if response["street_name"] else ""
            if "city" in response:
                organization.city = response["city"].strip() if response["city"] else ""
            if "province" in response:
                organization.province = response["province"].strip() if response["province"] else ""
            if "country" in response:
                organization.country = response["country"].strip() if response["country"] else ""
            if "postal_code" in response:
                organization.postal_code = response["postal_code"].strip() if response["postal_code"] else ""
            if "email" in response:
                organization.email = response["email"].strip() if response["email"] else ""
            if "phone" in response:
                phone = response["phone"].strip()
                phone = ''.join(i for i in phone if i.isdigit())
                organization.phone = phone
            if "website" in response:
                organization.website = response["website"].strip() if response["website"] else ""
            if "gender" in response:
                organization.gender_focus = ','.join(item for item in request.POST.getlist('gender'))
            if "age_group" in response:
                organization.age_group = ','.join(item for item in request.POST.getlist('age_group'))
            if "participants" in response:
                organization.participants = response["participants"].strip() if response["participants"] else ""
            if "sport_type" in response:
                save_organization_sports(profile, request.POST.getlist('sport_type'))
            save_organization_timings(profile, response)
            request.user.profile_status = True
            profile.profile_status = True
            profile.save()
            request.user.save()
            organization.save()
            context['organization'] = organization

            # messages.success(request, 'Organization details updated!')
            return home(request)
    return render(request, 'registration/organization_view.html', context)


def save_organization_timings(profile, response):
    if "sunday_start_time" in response and response["sunday_start_time"] != "" and "sunday_end_time" in response and \
            response["sunday_end_time"] != "":
        if Organization_Availability.objects.filter(profile=profile, day_of_week="Sunday").exists():
            obj = Organization_Availability.objects.get(profile=profile, day_of_week="Sunday")
            obj.start_time = response["sunday_start_time"]
            obj.end_time = response["sunday_end_time"]
            obj.save()
        else:
            obj = Organization_Availability(profile=profile, day_of_week="Sunday",
                                            start_time=response["sunday_start_time"],
                                            end_time=response["sunday_end_time"])
            obj.save()
    if "monday_start_time" in response and response["monday_start_time"] != "" and "monday_end_time" in response and \
            response["monday_end_time"] != "":
        if Organization_Availability.objects.filter(profile=profile, day_of_week="Monday").exists():
            obj = Organization_Availability.objects.get(profile=profile, day_of_week="Monday")
            obj.start_time = response["monday_start_time"]
            obj.end_time = response["monday_end_time"]
            obj.save()
        else:
            obj = Organization_Availability(profile=profile, day_of_week="Monday",
                                            start_time=response["monday_start_time"],
                                            end_time=response["monday_end_time"])
            obj.save()
    if "tuesday_start_time" in response and response["tuesday_start_time"] != "" and "tuesday_end_time" in response and \
            response["tuesday_end_time"] != "":
        if Organization_Availability.objects.filter(profile=profile, day_of_week="Tuesday").exists():
            obj = Organization_Availability.objects.get(profile=profile, day_of_week="Tuesday")
            obj.start_time = response["tuesday_start_time"]
            obj.end_time = response["tuesday_end_time"]
            obj.save()
        else:
            obj = Organization_Availability(profile=profile, day_of_week="Tuesday",
                                            start_time=response["tuesday_start_time"],
                                            end_time=response["tuesday_end_time"])
            obj.save()
    if "wednesday_start_time" in response and response[
        "wednesday_start_time"] != "" and "wednesday_end_time" in response and response["wednesday_end_time"] != "":
        if Organization_Availability.objects.filter(profile=profile, day_of_week="Wednesday").exists():
            obj = Organization_Availability.objects.get(profile=profile, day_of_week="Wednesday")
            obj.start_time = response["wednesday_start_time"]
            obj.end_time = response["wednesday_end_time"]
            obj.save()
        else:
            obj = Organization_Availability(profile=profile, day_of_week="Wednesday",
                                            start_time=response["wednesday_start_time"],
                                            end_time=response["wednesday_end_time"])
            obj.save()
    if "thursday_start_time" in response and response[
        "thursday_start_time"] != "" and "thursday_end_time" in response and response["thursday_end_time"] != "":
        if Organization_Availability.objects.filter(profile=profile, day_of_week="Thursday").exists():
            obj = Organization_Availability.objects.get(profile=profile, day_of_week="Thursday")
            obj.start_time = response["thursday_start_time"]
            obj.end_time = response["thursday_end_time"]
            obj.save()
        else:
            obj = Organization_Availability(profile=profile, day_of_week="Thursday",
                                            start_time=response["thursday_start_time"],
                                            end_time=response["thursday_end_time"])
            obj.save()
    if "friday_start_time" in response and response["friday_start_time"] != "" and "friday_end_time" in response and \
            response["friday_end_time"] != "":
        if Organization_Availability.objects.filter(profile=profile, day_of_week="Friday").exists():
            obj = Organization_Availability.objects.get(profile=profile, day_of_week="Friday")
            obj.start_time = response["friday_start_time"]
            obj.end_time = response["friday_end_time"]
            obj.save()
        else:
            obj = Organization_Availability(profile=profile, day_of_week="Friday",
                                            start_time=response["friday_start_time"],
                                            end_time=response["friday_end_time"])
            obj.save()
    if "saturday_start_time" in response and response[
        "saturday_start_time"] != "" and "saturday_end_time" in response and response["saturday_end_time"] != "":
        if Organization_Availability.objects.filter(profile=profile, day_of_week="Saturday").exists():
            obj = Organization_Availability.objects.get(profile=profile, day_of_week="Saturday")
            obj.start_time = response["saturday_start_time"]
            obj.end_time = response["saturday_end_time"]
            obj.save()
        else:
            obj = Organization_Availability(profile=profile, day_of_week="Saturday",

                                            start_time=response["saturday_start_time"],
                                            end_time=response["saturday_end_time"])
            obj.save()


def get_organization_timings(request):
    if request.is_ajax():
        profile = get_profile_from_user(request.user)
        availability = Organization_Availability.objects.filter(profile=profile)
        availability = list(availability.values("day_of_week", "start_time", "end_time", "pk"))
        return JsonResponse(availability, safe=False)


def save_organization_sports(profile, sport_type_list):
    user_choice = Secondary_SportsChoice.objects.filter(profile=profile)
    # Remove the sport choces from DB if they have been removed from current list
    if len(user_choice):
        for item in user_choice:
            if item.sport_type not in sport_type_list:
                Secondary_SportsChoice.objects.filter(id=item.pk).delete()

    # Add/Update sport choices for user
    for sport in sport_type_list:
        if not Secondary_SportsChoice.objects.filter(profile=profile, sport_type=sport).exists():
            obj = Secondary_SportsChoice(profile=profile, sport_type=sport)
            obj.save()
    return


def add_organization_locations(request):
    if request.method == "POST":
        selected_street = request.POST['selected_street_text'].strip() if request.POST['selected_street_text'] else ""
        selected_city = request.POST['selected_city_text'].strip() if request.POST['selected_city_text'] else ""
        selected_province = request.POST['selected_province_text'].strip() if request.POST[
            'selected_province_text'] else ""
        selected_country = request.POST['selected_country_text'].strip() if request.POST[
            'selected_country_text'] else ""
        selected_zipcode = request.POST['selected_zipcode_text'].strip() if request.POST[
            'selected_zipcode_text'] else ""

        try:
            profile = get_profile_from_user(request.user)
            if selected_city != "" and selected_province != "" and selected_country != "" and selected_zipcode != "":
                if Extra_Loctaions.objects.filter(profile=profile, street=selected_street, city=selected_city,
                                                  province=selected_province, country=selected_country,
                                                  zipcode=selected_zipcode).exists():
                    return JsonResponse({'status': 'Duplicate Location cannot be added!'}, safe=False)
                else:
                    obj = Extra_Loctaions(profile=profile, street=selected_street, city=selected_city,
                                          province=selected_province, country=selected_country,
                                          zipcode=selected_zipcode)
                    obj.save()
                return JsonResponse({'status': 'New Location added!'}, safe=False)
            else:
                return JsonResponse({'status': 'Missing values!'}, safe=False)
        except Exception:
            return JsonResponse({'status': 'An error occured!'}, safe=False)


def fetch_organization_locations(request):
    if request.is_ajax():
        profile = get_profile_from_user(request.user)
        location_choices = Extra_Loctaions.objects.filter(profile=profile).order_by("city")
        location_choices = list(location_choices.values("street", "city", "province", "country", "zipcode", "pk"))
        return JsonResponse(location_choices, safe=False)


def home(request):
    # Individual.objects.filter(pk=16).delete()
    # load_venues_excel()
    # load_pos_skill_type()

    if request.user.is_authenticated:
        if not request.user.profile_status:
            print('Redirecting to user profile...')
            return redirect('EventsApp:user_profile')

    sports = SportsType.objects.values('pk', 'sports_type_text').order_by('sports_type_text')

    advertisements = get_advertisements(request)
    advertisements = list(advertisements)
    # advertisements = [advertisements[i:i + 3] for i in range(0, len(advertisements), 3)]

    if request.user.is_authenticated:
        profile = get_profile_from_user(request.user)
    else:
        profile = None

    if request.user.is_authenticated and request.user.is_individual:

        user_sports = Secondary_SportsChoice.objects.filter(profile=profile).values('sport_type')
        for item in sports:
            flag = False
            for item2 in user_sports:
                if item['sports_type_text'] == item2['sport_type']:
                    flag = True

            if not flag:
                sports = sports.exclude(sports_type_text=item['sports_type_text'])

        user_loc = Extra_Loctaions.objects.filter(profile=profile).values_list('city', flat=True)
        venues = Venues.objects.values('pk', 'vm_name', 'vm_venuecity').filter(
            vm_venuecity__in=list(user_loc)).order_by('vm_name')
    else:
        venues = Venues.objects.values('pk', 'vm_name').order_by('vm_name')
    cities = Venues.objects.values('vm_venuecity').distinct().order_by('vm_venuecity')

    events = master_table.objects.all()
    events = get_events_by_time(events)

    recommended_events = []
    if request.user.is_authenticated:
        recommended_events = get_recommended_events(request)
        # recommended_events = master_table.objects.filter(pk__in=[event.pk for event in recommended_events])

    if request.GET.get('events_types'):
        selected_events_types = request.GET.getlist('events_types')
        if request.user.is_authenticated:
            for event in recommended_events[:]:
                if event.event_type not in selected_events_types:
                    recommended_events.remove(event)
        else:
            for event in events[:]:
                if event.event_type not in selected_events_types:
                    events.remove(event)

    if request.GET.get('sports'):
        selected_sports = request.GET.get('sports')
        if request.user.is_authenticated:
            for event in recommended_events[:]:
                if event.sport_type not in selected_sports:
                    recommended_events.remove(event)
        else:
            for event in events[:]:
                if event.sport_type not in selected_sports:
                    events.remove(event)

    if request.GET.get('cities'):
        selected_cities = request.GET.get('cities')
        if request.user.is_authenticated:
            for event in recommended_events[:]:
                if event.city not in selected_cities:
                    recommended_events.remove(event)
        else:
            for event in events[:]:
                if event.city not in selected_cities:
                    events.remove(event)

    if request.GET.get('venues'):
        selected_venues = request.GET.get('venues')
        if request.user.is_authenticated:
            for event in recommended_events[:]:
                if event.venue not in selected_venues:
                    recommended_events.remove(event)
        else:
            for event in events[:]:
                if event.venue not in selected_venues:
                    events.remove(event)

    if request.GET.get('date_range'):
        selected_date = request.GET.get('date_range')
        if selected_date != 'Select Date':
            selected_date = datetime.strptime(selected_date.strip(), '%Y-%m-%d').date()
            # print(selected_date)
            if request.user.is_authenticated:
                recommended_events = get_events_by_selected_date(recommended_events, selected_date)
            else:
                events = get_events_by_selected_date(events, selected_date)

    sort_events_by_date(events)

    if request.user.is_authenticated:
        sort_events_by_date(recommended_events)

    events = format_time(events)
    recommended_events = format_time(recommended_events)

    recommended_drop_in = []
    recommended_registrationList = []
    if request.user.is_authenticated:
        for event in recommended_events:
            sport_img = SportsImage.objects.filter(sport=event.sport_type).values("img")
            if len(sport_img):
                event.sport_logo = "/media/" + sport_img[0]["img"]
            else:
                event.sport_logo = "/media/images/Multisport.jpg"

        for event in recommended_events:
            if event.registration_type == "Registration":
                recommended_registrationList.append(event)
            elif event.registration_type == "Drop-in":
                recommended_drop_in.append(event)


        # recommended_registrationList = [recommended_registrationList[i:i + 3] for i in range(0, len(recommended_registrationList), 3)]
        # recommended_drop_in = [recommended_drop_in[i:i + 3] for i in range(0, len(recommended_drop_in), 3)]
        # recommended_events = [recommended_events[i:i + 3] for i in range(0, len(recommended_events), 3)]

    for event in events:
        sport_img = SportsImage.objects.filter(sport=event.sport_type).values("img")
        if len(sport_img):
            event.sport_logo = "/media/" + sport_img[0]["img"]
        else:
            event.sport_logo = "/media/images/Multisport.jpg"

    drop_in_eventList=[]
    registrationList=[]
    for event in events:
        if event.registration_type == "Registration":
            registrationList.append(event)
        elif event.registration_type == "Drop-in":
            drop_in_eventList.append(event)

    # registrationList = [registrationList[i:i + 3] for i in range(0, len(registrationList), 3)]
    # drop_in_eventList = [drop_in_eventList[i:i + 3] for i in range(0, len(drop_in_eventList), 3)]
    # events = [events[i:i + 3] for i in range(0, len(events), 3)]
    context = {
        'sports_list': sports,
        'venues_list': venues,
        'cities_list': cities,
        'registrationList': registrationList,
        'drop_in_eventList': drop_in_eventList,
        'recommended_registrationList': recommended_registrationList,
        'recommended_drop_in': recommended_drop_in,
        'advertisements': advertisements
    }
    # print(events)
    # html_template = loader.get_template('EventsApp/home.html')
    # return HttpResponse(html_template.render(context, request))

    return render(request, 'EventsApp/home.html', context)


def sort_events_by_date(events):
    def getDate(event):
        datetimes = event.datetimes if event.datetimes else event.current_datetimes
        time = datetimes.split("-")
        datetime_obj = datetime.strptime(time[0].strip(), '%m/%d/%Y %I:%M %p')
        return datetime_obj

    events.sort(key=lambda x: getDate(x), reverse=False)

    return


def get_events_by_time(events):
    events = list(events)
    full_events_list = []
    for event in events:
        # print(event.event_title)
        if event.datetimes:
            full_events_list.append(event)
        elif event.datetimes_all:
            all_dates_arr = event.datetimes_all.strip(',').split(',')
            for single_date in all_dates_arr:
                event_copy = copy.deepcopy(event)
                event_copy.current_datetimes = single_date
                full_events_list.append(event_copy)

    full_events_list.sort(key=lambda event: (event.datetimes if event.datetimes else event.current_datetimes))

    # for event in full_events_list:
    #     if event.datetimes:
    #         print("event.datetimes", event.event_title, event.datetimes)
    #     else:
    #         print("event.current_datetimes", event.event_title, event.current_datetimes)

    return full_events_list


def get_recommended_events(request):
    profile = get_profile_from_user(request.user)
    user_avaiability = Availability.objects.filter(profile=profile)

    events = master_table.objects.all()
    events = get_events_by_time(events)

    locations_saved = Extra_Loctaions.objects.filter(profile=profile)
    loc_list = [item.city.lower() for item in locations_saved]
    recommended_events = []

    # FILTER by DateTime
    # If there is no user availability then add all events to recommended events
    if len(user_avaiability):
        for event in events:
            event_date, event_start_time, event_end_time = extract_event_datetime(event)
            for avail in user_avaiability:
                if avail.day_of_week == (event_date.weekday() + 1):
                    if event_start_time >= avail.start_time and event_end_time <= avail.end_time:
                        recommended_events.append(event)
                        # print("Added", event.event_title, event_date, event_start_time, event_end_time)
                        break
    else:
        recommended_events = [event for event in events]

    # print("Availability Filter", recommended_events)

    # FILTER BY Location
    for event in recommended_events[:]:
        if event.city is not None:
            if event.city.lower() not in loc_list:
                recommended_events.remove(event)

    recommended_events = list(recommended_events)
    # print("Location Filter", recommended_events)

    if request.user.is_individual:
        individual = Individual.objects.get(profile=profile)

        # FILTER BY Age
        if individual.dob:
            dob = datetime.strptime(individual.dob, '%Y-%m-%d').date()
            today = date.today()
            age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
            for event in recommended_events[:]:
                positions = Events_PositionInfo.objects.filter(event=event)
                age_fail_count = 0
                for position in positions:
                    if position.max_age >= age >= position.min_age:
                        continue
                    else:
                        age_fail_count = age_fail_count + 1

                if age_fail_count == len(positions):
                    recommended_events.remove(event)

    # print("Age Filter", recommended_events)

    # FILTER BY Gender
    if request.user.is_individual:
        individual = Individual.objects.get(profile=profile)

        individual_gender = []
        if individual.participation_interest and individual.participation_interest != "":
            # print(individual.participation_interest)
            individual_gender = individual.participation_interest.split(",")
        # print(individual_gender)
        for event in recommended_events[:]:
            flag = 0
            if event.gender:
                gender_list = event.gender.split(",")
                for item in individual_gender:
                    if item in gender_list:
                        flag = flag + 1
                        # print(event.event_title, gender_list, individual_gender, flag, item)

                if flag == 0:
                    # print(event.event_title, gender_list, individual_gender, flag)
                    recommended_events.remove(event)

    # print("Gender Filter", recommended_events)

    # FILTER BY Sports
    sport_choices = Secondary_SportsChoice.objects.filter(profile=profile).order_by("sport_type")
    sports_list = []
    for item in sport_choices:
        sports_list.append(item.sport_type)

    for event in recommended_events[:]:
        if event.sport_type not in sports_list:
            recommended_events.remove(event)

    # print("Sports Filter", recommended_events)

    # FILTER BY Positions
    if request.user.is_individual:
        position_choices = Secondary_SportsChoice.objects.filter(profile=profile)
        position_list = []
        for item in position_choices:
            position_list.append(item.position)

        for event in recommended_events[:]:
            event_positions_list = Events_PositionInfo.objects.filter(event=event.pk)
            flag = 0
            for item in event_positions_list:
                if item.position_name not in position_list:
                    flag = flag + 1

            if flag > 0:
                recommended_events.remove(event)

    # print("Positions Filter", recommended_events)

    # FILTER BY Skills
    if request.user.is_individual:
        skill_choices = Secondary_SportsChoice.objects.filter(profile=profile)

        skill_list = []
        for item in skill_choices:
            skill_list.append(item.skill)

        for event in recommended_events[:]:
            event_skill_list = Events_PositionInfo.objects.filter(event=event.pk)
            flag = 0
            for item in event_skill_list:
                if item.position_type not in skill_list:
                    flag = flag + 1

            if flag > 0:
                recommended_events.remove(event)

    # print("Skills Filter", recommended_events)

    return recommended_events


def extract_event_datetime(event):
    event_date = ""
    event_start_time = ""
    event_end_time = ""

    times_list=[]

    if event.datetimes:
        split_list = event.datetimes.split("-")
        date1 = split_list[0]
        date2 = split_list[-1]
        times_list.append(date1)
        times_list.append(date2)
    elif event.current_datetimes:
        split_list = event.current_datetimes.split("-")
        date1 = split_list[0]
        date2 = event.current_datetimes[0:10] + split_list[1]
        times_list.append(date1)
        times_list.append(date2)

    if len(times_list) > 0:
        event_start_datetime = datetime.strptime(times_list[0].strip(), '%m/%d/%Y %I:%M %p')
        event_end_datetime = datetime.strptime(times_list[-1].strip(), '%m/%d/%Y %I:%M %p')
        event_date = event_start_datetime.date()
        event_start_time = event_start_datetime.time()
        event_end_time = event_end_datetime.time()

        # print(event_date, event_start_time, event_end_time)

        return event_date, event_start_time, event_end_time

    return event_date, event_start_time, event_end_time


def format_time(events):
    # for event in events:
    #     if event.datetimes:
    #         print("event.datetimes", event.event_title, event.datetimes)
    #     else:
    #         print("event.current_datetimes", event.event_title, event.current_datetimes)

    for event in events:
        if event.datetimes:
            time = event.datetimes.split("-")
            str_datetime = format_time_helper(time)
            event.datetimes = str_datetime
        elif event.current_datetimes:
            string_date = datetime.strptime(event.current_datetimes[0:10], '%m/%d/%Y').date()
            str_datetime = string_date.strftime("%B %d") + " from " + event.current_datetimes[
                                                                      10:len(event.current_datetimes)]
            event.current_datetimes = str_datetime

    return events


def format_time_helper(time):
    if time is not None and time != "":
        start_time = datetime.strptime(time[0].strip(), '%m/%d/%Y %I:%M %p').time()
        end_time = datetime.strptime(time[-1].strip(), '%m/%d/%Y %I:%M %p').time()
        start_time = start_time.strftime("%I:%M %p")
        end_time = end_time.strftime('%I:%M %p')

        start_date = datetime.strptime(time[0].strip(), '%m/%d/%Y %I:%M %p').date()
        end_date = datetime.strptime(time[-1].strip(), '%m/%d/%Y %I:%M %p').date()
        start_date = start_date.strftime("%B %d")
        end_date = end_date.strftime("%B %d")
        if start_date == end_date:
            str_datetime = start_date + " from " + start_time + " - " + end_time
        else:
            str_datetime = start_date + " - " + end_date + " from " + start_time + " - " + end_time

    return str_datetime


def get_events_by_selected_date(events_list, selected_date):
    new_events_list=[]
    for event in events_list[:]:
        if event.current_datetimes:
            evt_date = event.current_datetimes[0:10]
            evt_date = datetime.strptime(evt_date.strip(), '%m/%d/%Y').date()
            if selected_date == evt_date:
                new_events_list.append(event)

    return new_events_list


@login_required
def event_details(request, event_id, event_date):
    context = {}
    event = master_table.objects.get(pk=event_id)
    profile = get_profile_from_user(request.user)
    event_postions = list(Events_PositionInfo.objects.filter(event=event_id))
    # print(event_postions)
    if event.is_recurring:
        for evnt_pos in event_postions[:]:
            string_date = datetime.strptime(evnt_pos.datetimes[0:10], '%m/%d/%Y').date()
            str_datetime = string_date.strftime("%B %d") + " from " + evnt_pos.datetimes[10:len(evnt_pos.datetimes)]
            if str_datetime != event_date:
                event_postions.remove(evnt_pos)

    context['event'] = event
    context['event_postions'] = event_postions
    context['event_date'] = event_date

    if request.method == 'POST':
        # print(request.POST.dict())
        response = request.POST.dict()
        for key in response:
            if 'chk' in key:
                print(key)
                idx = key.split("_")[-1]
                position_id = response['posId_' + idx]
                pos_name = response['posName_' + idx]
                skill = response['skill_' + idx]
                needed_pos = response['needed_' + idx]
                no_of_pos = response['noOfPos_' + idx]
                pos_cost = response['cost_' + idx]
                ## Fetch previous unpurchased orderitems to set checkout timer to current
                remove_expired_cart_items(request)
                current_cart = OrderItems.objects.filter(profile=profile, purchased=False)
                for current_item in current_cart:
                    current_item.checkout_timer = datetime.now(timezone.utc)
                    current_item.save()
                ## Add to cart
                cart = OrderItems()
                cart.event = master_table.objects.get(pk=event_id)
                cart.profile = profile
                cart.position_id = Events_PositionInfo.objects.get(pk=position_id)
                cart.date = date.today()
                cart.position_type = pos_name
                cart.skill = skill
                cart.no_of_position = needed_pos
                cart.position_cost = pos_cost
                cart.total_cost = float(pos_cost) * int(needed_pos)
                cart.checkout_timer = datetime.now(timezone.utc)
                cart.save()

                # Update Inventory
                event_pos = Events_PositionInfo.objects.get(pk=position_id)
                event_pos.no_of_position = int(event_pos.no_of_position) - int(needed_pos)
                event_pos.save()

                # print(request.user)

                if request.user.is_individual:
                    sub_prof = Individual.objects.get(profile=profile)
                elif request.user.is_organization:
                    sub_prof = Organization.objects.get(profile=profile)

                # Email Creator - New Subscriber
                event_subject = "New subscriber for Event: " + event.event_title
                event_message = "A new user has subscribed to event: " + event.event_title + "\n" + \
                                "Subscriber Name: " + sub_prof.first_name + " " + \
                                sub_prof.last_name if sub_prof.last_name else "" + "\n" + \
                                                                              "Subscriber Email: " + request.user.email + "\n"
                # if event.created_by:
                #     util.email(event_subject, event_message, [event.created_by.email])

                messages.info(request, "Order Items will be present in your cart for only 30 mins!")

        return redirect('EventsApp:cart_summary')

    return render(request, "EventsApp/detail_dashboard.html", context)


@csrf_exempt
def delete_cart_item(request):
    if request.POST:
        try:
            event_pos_id = request.POST['event_pos_id']
            cart_item_id = request.POST['cart_item_id']
            # print(event_pos_id, cart_item_id)
            order_item = OrderItems.objects.get(pk=cart_item_id)
            evnt_pos_info = Events_PositionInfo.objects.get(pk=event_pos_id)
            evnt_pos_info.no_of_position = evnt_pos_info.no_of_position + order_item.no_of_position
            evnt_pos_info.save()
            order_item.delete()
            return JsonResponse({'status': 'Order Item deleted!'}, safe=False)
        except:
            return JsonResponse({'status': 'Some error occurred!'}, safe=False)


def fetch_cart_items(request):
    total = 0
    order_items = []
    remove_expired_cart_items(request)
    profile = get_profile_from_user(request.user)
    if request.is_ajax():
        cart = OrderItems.objects.filter(profile=profile, purchased=False)
        for item in cart:
            total = total + item.total_cost

        for item in cart:
            order_items.append({
                'event_title': item.event.event_title,
                'position_type': item.position_type,
                'no_of_position': item.no_of_position,
                'position_cost': item.position_cost,
                'position_id': item.position_id.pk,
                'pk': item.pk,
            })
        return JsonResponse({"cart": order_items, "total": total}, safe=False)


def remove_expired_cart_items(request):
    profile = get_profile_from_user(request.user)
    cart = OrderItems.objects.filter(profile=profile, purchased=False)
    for item in cart:
        time_delta = (datetime.now(timezone.utc) - item.checkout_timer)
        total_mins = (time_delta.total_seconds()) / 60
        if total_mins > 30:
            order_item = OrderItems.objects.get(pk=item.pk)
            evnt_pos_info = Events_PositionInfo.objects.get(pk=item.position_id.pk)
            evnt_pos_info.no_of_position = evnt_pos_info.no_of_position + order_item.no_of_position
            evnt_pos_info.save()
            order_item.delete()


@login_required
def cart_summary(request):
    context = {}
    total = 0
    profile = get_profile_from_user(request.user)
    remove_expired_cart_items(request)
    context['STRIPE_PUBLISHABLE_KEY'] = settings.STRIPE_PUBLISHABLE_KEY,
    cart = OrderItems.objects.filter(profile=profile, purchased=False)
    for item in cart:
        total = total + item.total_cost

    context["cart"] = cart
    context["total"] = total
    if request.method == 'POST':
        order_obj = Order.objects.filter(customer=profile, payment=False)
        if order_obj.exists():
            order_obj = Order.objects.get(customer=profile, payment=False)
            order_obj.items.clear()
            order_obj.order_date = timezone.now()
            for order_items in cart:
                order_obj.items.add(order_items)
                total = total + order_items.total_cost

            order_obj.order_amount = total
        else:
            order = Order()
            order.customer = profile
            order.order_date = timezone.now()
            order.order_amount = total
            order.payment = False
            order.save()
            for order_items in cart:
                order.items.add(order_items)

        return redirect('EventsApp:charge')

    return render(request, "EventsApp/cart_summary.html", context)


@login_required
def charge(request):
    context = {}
    char_set = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
    key = settings.STRIPE_PUBLISHABLE_KEY
    context['key'] = key
    order = Order.objects.get(customer=Profile.objects.get(active_user=request.user), payment=False)
    cart = order.items.all()
    context["order"] = order
    context["cart"] = cart
    context['orderAmount'] = order.order_amount
    totalCents = int(order.order_amount * 100)
    context['totalCents'] = totalCents
    print(cart)
    if request.method == 'POST':
        try:
            charge = stripe.Charge.create(amount=totalCents,
                                          currency='cad',
                                          description=order,
                                          source=request.POST['stripeToken'])
            print(charge)
            if charge.status == "succeeded":
                print('payment success')
                orderId = get_random_string(length=16, allowed_chars=char_set)
                paymentId = get_random_string(length=16, allowed_chars=char_set)
                order.orderId = f'#{request.user}{orderId}'
                order.paymentId = paymentId
                order.payment = True
                order.payment_method = "Card"
                for order_item in order.items.all():
                    item = OrderItems.objects.get(pk=order_item.pk)
                    item.purchased = True
                    item.save()

                order.save()

                return redirect('EventsApp:payment-success')

            elif charge.status == "failed":
                return redirect('EventsApp:payment-cancel')

        except Exception as e:
            print(e)

    return render(request, 'EventsApp/charge.html', context)


def paymentSuccess(request):
    print("Success ho gaya")
    context = {
        "payment_status": "success"
    }
    return render(request, "EventsApp/confirmation.html", context)


def paymentCancel(request):
    print("cancel ho gaya")
    context = {
        "payment_status": "fail"
    }
    return render(request, "EventsApp/confirmation.html", context)


def pay_at_venue(request):
    context = {
        "payment_status": "VenuePay"
    }
    profile = get_profile_from_user(request.user)
    char_set = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
    order = Order.objects.get(customer=profile, payment=False)
    print(order.order_amount)
    orderId = get_random_string(length=16, allowed_chars=char_set)
    paymentId = get_random_string(length=16, allowed_chars=char_set)
    order.orderId = f'#{request.user}{orderId}'
    order.paymentId = paymentId
    order.payment = True
    order.payment_method = "Cash Pickup"
    for order_item in order.items.all():
        item = OrderItems.objects.get(pk=order_item.pk)
        item.purchased = True
        item.save()

    order.save()

    # order_subject = "Order Confirmed: " + f'#{request.user}{orderId}'
    # order_message = "Your seat has been reserved. Payment would be made on the day of the event. " + \
    #                 "Customer Name: " + request.user.first_name + " " + request.user.last_name + "\n" + \
    #                 "Customer Email: " + request.user.email + "\n"
    #
    # util.email(order_subject, order_message, [request.user.email])

    return render(request, "EventsApp/confirmation.html", context)


@login_required
def add_availability(request):
    context = {}
    profile = get_profile_from_user(request.user)
    form = AvailabilityForm(request.POST or None,
                            instance=Availability(),
                            initial={'profile': profile})

    context['form'] = form
    # user = User.objects.get(email=request.user.email)
    user_avaiability = Availability.objects.filter(profile=profile)
    get_day_of_week(user_avaiability)

    context["user_availability"] = user_avaiability
    if request.POST:
        if form.is_valid():
            obj = Availability(profile=profile,
                               day_of_week=form.cleaned_data['day_of_week'],
                               start_time=form.cleaned_data['start_time'],
                               end_time=form.cleaned_data['end_time'],
                               all_day=form.cleaned_data['all_day'])
            is_duplicate = check_duplicate_availability(user_avaiability, obj)
            if is_duplicate:
                messages.error(request, "Duplicate Availability!")
            else:
                obj.save()
                user_avaiability = Availability.objects.filter(profile=profile)
                get_day_of_week(user_avaiability)
                context["user_availability"] = user_avaiability
                messages.success(request, "New Availability Added!")
        else:
            messages.error(request, "Enter a valid time!")
            print(form.errors)

    form = AvailabilityForm()
    context['form'] = form
    return render(request, "EventsApp/add_availability.html", context)


def get_user_availability(request):
    profile = get_profile_from_user(request.user)
    if request.is_ajax():
        # user = User.objects.get(email=request.user.email)
        user_availability = Availability.objects.filter(profile=profile)
        user_availability = list(user_availability.values("day_of_week", "start_time", "end_time", "pk", "all_day"))
        for avail in user_availability:
            if avail["day_of_week"] == 1:
                avail["day_of_week"] = "Monday"
            if avail["day_of_week"] == 2:
                avail["day_of_week"] = "Tuesday"
            if avail["day_of_week"] == 3:
                avail["day_of_week"] = "Wednesday"
            if avail["day_of_week"] == 4:
                avail["day_of_week"] = "Thursday"
            if avail["day_of_week"] == 5:
                avail["day_of_week"] = "Friday"
            if avail["day_of_week"] == 6:
                avail["day_of_week"] = "Saturday"
            if avail["day_of_week"] == 7:
                avail["day_of_week"] = "Sunday"

            avail["start_time"] = avail["start_time"].strftime("%I:%M %p")
            avail["end_time"] = avail["end_time"].strftime("%I:%M %p")

    return JsonResponse(user_availability, safe=False)


def add_user_availability(request):
    # user = User.objects.get(email=request.user.email)
    profile = get_profile_from_user(request.user)
    user_avaiability = Availability.objects.filter(profile=profile)
    if request.method == "POST":
        obj = Availability(profile=profile,
                           day_of_week=request.POST['day_of_week'],
                           start_time=request.POST['start_time'],
                           end_time=request.POST['end_time'],
                           all_day=request.POST['all_day'] == "true")
        is_duplicate = check_duplicate_availability(user_avaiability, obj)
        if is_duplicate:
            return JsonResponse({'status': 'Duplicate Availability!'}, safe=False)
        else:
            obj.save()
            user_avaiability = Availability.objects.filter(profile=profile)
            get_day_of_week(user_avaiability)
            return JsonResponse({'status': 'New Availability Added!'}, safe=False)
    else:
        return JsonResponse({'status': 'Enter a valid time!'}, safe=False)


def check_duplicate_availability(user_avaiability, new_availability):
    if new_availability.day_of_week == 1:
        new_availability.day_of_week = "Monday"
    if new_availability.day_of_week == 2:
        new_availability.day_of_week = "Tuesday"
    if new_availability.day_of_week == 3:
        new_availability.day_of_week = "Wednesday"
    if new_availability.day_of_week == 4:
        new_availability.day_of_week = "Thursday"
    if new_availability.day_of_week == 5:
        new_availability.day_of_week = "Friday"
    if new_availability.day_of_week == 6:
        new_availability.day_of_week = "Saturday"
    if new_availability.day_of_week == 7:
        new_availability.day_of_week = "Sunday"

    for avail in user_avaiability:
        if avail.day_of_week == int(new_availability.day_of_week) and \
                (new_availability.start_time >= avail.start_time.strftime("%H:%M") or \
                 new_availability.end_time <= avail.end_time.strftime("%H:%M")):
            return True

    return False


@login_required
def delete_availability(request):
    if request.POST:
        try:
            id = request.POST['id']
            avail = Availability.objects.get(pk=id)
            avail.delete()
            return JsonResponse({'status': 'Availability removed successfully!'}, safe=False)
        except:
            return JsonResponse({'status': 'Some error occurred!'}, safe=False)


def get_day_of_week(user_avaiability):
    for avail in user_avaiability:
        if avail.day_of_week == 1:
            avail.day_of_week = "Monday"
        if avail.day_of_week == 2:
            avail.day_of_week = "Tuesday"
        if avail.day_of_week == 3:
            avail.day_of_week = "Wednesday"
        if avail.day_of_week == 4:
            avail.day_of_week = "Thursday"
        if avail.day_of_week == 5:
            avail.day_of_week = "Friday"
        if avail.day_of_week == 6:
            avail.day_of_week = "Saturday"
        if avail.day_of_week == 7:
            avail.day_of_week = "Sunday"


def load_venues_excel():
    path = "./latest_venue.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    # Venues.objects.all().delete()
    for i in range(2, sheet_obj.max_row + 1):
        if sheet_obj.cell(row=i, column=5).value.strip() == "ON" or sheet_obj.cell(row=i,
                                                                                   column=5).value.strip() == "Ontario":
            vm_name = sheet_obj.cell(row=i, column=1).value.strip()
            vm_venue_description = sheet_obj.cell(row=i, column=2).value.strip()
            vm_venue_street = sheet_obj.cell(row=i, column=3).value.strip()
            vm_venuecity = sheet_obj.cell(row=i, column=4).value.strip()
            vm_venue_province = sheet_obj.cell(row=i, column=5).value.strip()
            vm_venue_country = sheet_obj.cell(row=i, column=6).value.strip()
            vm_venue_zip = sheet_obj.cell(row=i, column=7).value.strip()
            if not Venues.objects.filter(vm_name=vm_name, vm_venue_description=vm_venue_description, vm_venue_street=vm_venue_street,
                         vm_venuecity=vm_venuecity, vm_venue_province=vm_venue_province,
                         vm_venue_country=vm_venue_country, vm_venue_zip=vm_venue_zip).exists():
                obj = Venues(vm_name=vm_name, vm_venue_description=vm_venue_description, vm_venue_street=vm_venue_street,
                             vm_venuecity=vm_venuecity, vm_venue_province=vm_venue_province,
                             vm_venue_country=vm_venue_country, vm_venue_zip=vm_venue_zip)
                print(i)
                obj.save()


def load_pos_skill_type():
    path = "./sports_db.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    for i in range(2, sheet_obj.max_row + 1):
        sports_catgeory_text = sheet_obj.cell(row=i, column=1).value.strip()
        if not SportsCategory.objects.filter(sports_catgeory_text=sports_catgeory_text).exists():
            sc = SportsCategory(sports_catgeory_text=sports_catgeory_text)
            sc.save()

        sports_category = SportsCategory.objects.get(sports_catgeory_text=sports_catgeory_text)
        sports_type_text = sheet_obj.cell(row=i, column=2).value.strip()

        if not SportsType.objects.filter(sports_category=sports_category, sports_type_text=sports_type_text).exists():
            st = SportsType(sports_category=sports_category, sports_type_text=sports_type_text)
            st.save()

        sports_type = SportsType.objects.get(sports_type_text=sports_type_text)

        position = sheet_obj.cell(row=i, column=3).value.strip()
        skill = sheet_obj.cell(row=i, column=4).value.strip()
        if not PositionAndSkillType.objects.filter(sports_category=sports_category,
                                   sports_type=sports_type,
                                   position_type=position,
                                   skill_type=skill).exists():
            obj = PositionAndSkillType(sports_category=sports_category,
                                       sports_type=sports_type,
                                       position_type=position,
                                       skill_type=skill)
            print(i, sports_category, sports_type, obj)
            obj.save()


@login_required
def delete_by_id(request, event_id, date):
    try:
        updated_dates = ""
        month = list(calendar.month_name).index(date.split(' ')[0])
        day = date.split(' ')[1]
        date = str(month) + '/' + day
        event = master_table.objects.get(pk=event_id)
        if event.is_recurring:
            for existing_dates in event.datetimes_all.split(','):
                if date not in existing_dates:
                    updated_dates += existing_dates + ","
            if updated_dates.strip(',') == "":
                event.delete()
            else:
                event.datetimes_all = updated_dates.strip(',')
                event.save()
        else:
            event.delete()
        user = User.objects.get(email=request.user.email)
        event_subject = "Event Cancelled - " + event.event_title
        event_data = " Event Title: " + event.event_title + "\n Event Description: " \
                     + event.description + "\n Event Location: " + event.venue + ", " + event.city \
                     + "\n Event Dates: " + \
                     (event.datetimes if not event.is_recurring else date)

        for cart_item in OrderItems.objects.filter(event=event):
            subs_email = cart_item.user.email
            util.email(event_subject, "Hello " + user.first_name + " " + user.last_name
                       + ", the following subscribed event in your cart has been cancelled by the creator:\n\n"
                       + event_data
                       , [subs_email])
        for order_item in Order.objects.filter(event=event):
            order_email = order_item.customer.email
            util.email(event_subject, "Hello " + user.first_name + " " + user.last_name
                       + ", the following event has been cancelled by the creator:\n\n" + event_data
                       , [order_email])
        util.email(event_subject, "Hello " + user.first_name + " " + user.last_name
                   + ", the following event has been cancelled as requested (subscribers and customers have been "
                     "notified):\n\n" + event_data
                   , [user.email])
        messages.success(request, "Event removed successfully!")

    except Exception as e:
        print(e)

    return redirect('EventsApp:list-events')


@login_required
def invite_by_id(request, event_id, email=None):
    ## only works for individual
    context = {}
    profile = get_profile_from_user(request.user)
    form = InviteForm(request.POST or None,
                      instance=Invite(),
                      initial={'profile': profile})
    individual = Individual.objects.get(profile=profile)
    context['form'] = form
    # user = User.objects.get(email=request.user.email)
    event = master_table.objects.get(pk=event_id)
    context["event"] = event
    context["invites"] = Invite.objects.all().filter(event=event).distinct('email')

    if email:
        if individual.first_name and individual.last_name:
            full_name = individual.first_name + " " + individual.last_name
        else:
            full_name = ""
        util.email("Invitation from Insportify", "Hi there! " + full_name
                   + " has invited you to the event: " + event.event_title
                   + ". Join Insportify now: " + request.get_host() + "/" + str(event_id),
                   [email])
        messages.success(request, "Another invitation email sent to " + email)

    if request.POST:
        if form.is_valid():
            obj = Invite(profile=profile,
                         event=event,
                         email=form.cleaned_data['email'])
            obj.save()
            if individual.first_name and individual.last_name:
                full_name = individual.first_name + " " + individual.last_name
            else:
                full_name = ""
            util.email("Invitation from Insportify", "Hi there! " + full_name
                       + " has invited you to the event: " + event.event_title
                       + ". Join Insportify now: " + request.get_host() + "/" + str(event_id),
                       [form.cleaned_data['email']])
            messages.success(request, "Invitation email sent to " + form.cleaned_data['email'])
            context["invites"] = Invite.objects.all().filter(event=event).distinct('email')
        else:
            messages.error(request, form.errors)

    return render(request, "EventsApp/invite.html", context)


@login_required
def invite(request):
    ## works only for individual
    context = {}
    profile = get_profile_from_user(request.user)
    form = InviteForm(request.POST or None,
                      instance=Invite(),
                      initial={'profile': profile})

    context['form'] = form
    # user = User.objects.get(email=request.user.email)
    context["invites"] = Invite.objects.all().filter(profile=profile)
    individual = Individual.objects.get(profile=profile)
    if request.POST:
        if form.is_valid():
            obj = Invite(profile=profile,
                         event=None,
                         email=form.cleaned_data['email'])
            obj.save()
            if individual.first_name and individual.last_name:
                full_name = individual.first_name + " " + individual.last_name
            else:
                full_name = ""
            util.email("Invitation from Insportify", "Hi there! " + full_name
                       + " has invited you to Insportify! Join Now: " + "http://127.0.0.1:8000/",
                       [form.cleaned_data['email']])
            context["invites"] = Invite.objects.all().filter(profile=profile)
        else:
            print(form.errors)

    return render(request, "EventsApp/invite.html", context)


@login_required
def logo_upload_view(request):
    profile = get_profile_from_user(request.user)
    if request.method == 'POST':
        img_obj = ""
        form = LogoForm(request.POST, request.FILES)
        if form.is_valid():
            if Logo.objects.filter(profile=profile).exists():
                img_obj = Logo.objects.get(profile=profile)
                img_obj.image = form.instance.image
                img_obj.save()
            else:
                obj = form.save(commit=False)
                obj.profile = profile
                obj.save()
            return render(request, 'EventsApp/add_logo.html', {'form': form, 'img_obj': img_obj})
    else:
        img_obj = ""
        form = LogoForm()
        if Logo.objects.filter(profile=profile).exists():
            img_obj = Logo.objects.get(profile=profile)

    return render(request, 'EventsApp/add_logo.html', {'form': form, 'img_obj': img_obj})


def days_between(start, end, week_day, start_time, end_time):
    first_day = datetime.strptime(start, '%m/%d/%Y')
    last_day = datetime.strptime(end, '%m/%d/%Y')
    dates = [first_day + timedelta(days=x) for x in range((last_day - first_day).days + 1) if
             (first_day + timedelta(days=x)).weekday() == time.strptime(week_day, '%A').tm_wday]
    return [date_obj.strftime('%m/%d/%Y') + ' ' + start_time + ' - ' + end_time for date_obj in dates]


def remove_exceptions_from_recurring_days(all_dates, exception_dates):
    all_dates_arr = all_dates.strip(',').split(',')
    exception_dates_arr = exception_dates.split(',')
    print(all_dates_arr)
    print(exception_dates_arr)
    filtered_dates = all_dates.strip(',').split(',')
    for ex_date in exception_dates_arr:
        for date in all_dates_arr:
            if ex_date == date:
                filtered_dates.remove(date)
    filtered_dates.sort()
    # print(filtered_dates)
    return ','.join(filtered_dates)


def get_advertisements(request):
    ads = []
    if not request.user.is_authenticated:
        ads = Advertisement.objects.all().filter(Q(end_time__gte=date.today()) & Q(geographical_scope="National"))
    else:
        profile = get_profile_from_user(request.user)
        locs = Extra_Loctaions.objects.all().filter(profile=profile)
        city_str = ''
        prov_str = ''
        for loc in locs:
            city_str += loc.city + ' '
            prov_str += loc.province + ' '
        ads = Advertisement.objects.all().filter(Q(end_time__gte=date.today()) & Q(geographical_scope="National") |
                                                 (Q(geographical_scope="Provincial") & Q(
                                                     province__icontains=prov_str)) |
                                                 (Q(geographical_scope="Local") & Q(city__icontains=city_str)))
    return ads


def show_advertisement(request, header):
    try:
        advert = Advertisement.objects.get(header=header)
        if advert is not None:
            if advert.hit_count is None:
                advert.hit_count = 1
            else:
                advert.hit_count += 1
            advert.save()
            # print(advert.url)
            return redirect("https://" + advert.url)
    except:
        return home(request)
    return home(request)



def sports_type_excel():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SPORTS DB"
    sheet["A1"] = "Sports Category"
    sheet["B1"] = "Sports Type"
    sheet["C1"] = "Position"
    sheet["D1"] = "Skill"

    pos_skill = PositionAndSkillType.objects.all()
    row=2
    for item in pos_skill:
        print(item)
        c1 = sheet.cell(row=row, column=1)
        c1.value = item.sports_category.sports_catgeory_text

        c2 = sheet.cell(row=row, column=2)
        c2.value = item.sports_type.sports_type_text

        c3 = sheet.cell(row=row, column=3)
        c3.value = item.position_type

        c4 = sheet.cell(row=row, column=4)
        c4.value = item.skill_type

        row=row+1

    workbook.save(filename='sports_db.xlsx')


def venue_excel():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Latest Venue List"
    sheet["A1"] = "vm_name"
    sheet["B1"] = "vm_venue_description"
    sheet["C1"] = "vm_venue_street"
    sheet["D1"] = "vm_venuecity"
    sheet["E1"] = "vm_venue_province"
    sheet["F1"] = "vm_venue_country"
    sheet["G1"] = "vm_venue_zip"

    venues = Venues.objects.all()
    row=2
    for item in venues:
        print(item)
        c1 = sheet.cell(row=row, column=1)
        c1.value = item.vm_name

        c2 = sheet.cell(row=row, column=2)
        c2.value = item.vm_venue_description

        c3 = sheet.cell(row=row, column=3)
        c3.value = item.vm_venue_street

        c4 = sheet.cell(row=row, column=4)
        c4.value = item.vm_venuecity

        c5 = sheet.cell(row=row, column=5)
        c5.value = item.vm_venue_province

        c6 = sheet.cell(row=row, column=6)
        c6.value = item.vm_venue_country

        c7 = sheet.cell(row=row, column=7)
        c7.value = item.vm_venue_zip

        row=row+1

    workbook.save(filename='ON_latest_venue.xlsx')